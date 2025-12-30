"""
Testy jednostkowe dla Dziennika Kolarza.

Testuje tworzenie arkuszy z metrykami WKO5/INSCYD.
"""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from kombajn.config import (
    DEFAULTS,
    POWER_DEFAULTS,
    METABOLIC_DEFAULTS,
    SHEET_CONFIG,
    COLORS,
    LOG_HEADERS,
    CHO_HEADERS,
    POWER_ZONES,
)
from kombajn.styles import ExcelStyles, DEFAULT_STYLES
from kombajn.utils import (
    sanitize_filename,
    validate_output_path,
    safe_save_workbook,
)
from kombajn.sheets import (
    SettingsSheet,
    LogSheet,
    DashboardSheet,
    CHOSourcesSheet,
    PowerZonesSheet,
)
from kombajn.main import create_workbook


class TestConfig:
    """Testy konfiguracji."""
    
    def test_defaults_values(self):
        """Sprawdza wartości domyślne."""
        assert DEFAULTS.BMR == 1800
        assert DEFAULTS.TEF == 200
        assert DEFAULTS.NEAT == 300
    
    def test_power_defaults(self):
        """Sprawdza domyślne wartości mocy (WKO5)."""
        assert POWER_DEFAULTS.FTP == 250
        assert POWER_DEFAULTS.MAX_HR == 185
        assert POWER_DEFAULTS.WEIGHT_KG == 75.0
    
    def test_metabolic_defaults(self):
        """Sprawdza domyślne wartości metaboliczne (INSCYD)."""
        assert METABOLIC_DEFAULTS.VO2MAX == 55.0
        assert METABOLIC_DEFAULTS.VLAMAX == 0.50
    
    def test_power_zones_count(self):
        """Sprawdza liczbę stref mocy."""
        assert len(POWER_ZONES) == 7
    
    def test_power_zones_structure(self):
        """Sprawdza strukturę stref mocy."""
        zone1 = POWER_ZONES[0]
        assert zone1.number == 1
        assert zone1.min_pct == 0.0
        assert zone1.max_pct == 0.55
    
    def test_sheet_config(self):
        """Sprawdza konfigurację arkuszy."""
        assert SHEET_CONFIG.MAX_LOG_ROWS == 1000
        assert SHEET_CONFIG.CTL_DAYS == 42
        assert SHEET_CONFIG.ATL_DAYS == 7
    
    def test_log_headers_count(self):
        """Sprawdza liczbę nagłówków dziennika (42 kolumny)."""
        assert len(LOG_HEADERS) == 42
    
    def test_log_headers_contain_power_metrics(self):
        """Sprawdza czy nagłówki zawierają metryki mocy."""
        assert "NP (W)" in LOG_HEADERS
        assert "IF" in LOG_HEADERS
        assert "TSS" in LOG_HEADERS
        assert "CTL" in LOG_HEADERS
        assert "ATL" in LOG_HEADERS
        assert "TSB" in LOG_HEADERS


class TestStyles:
    """Testy stylów Excel."""
    
    def test_default_styles_exist(self):
        """Sprawdza istnienie domyślnych stylów."""
        assert DEFAULT_STYLES is not None
        assert isinstance(DEFAULT_STYLES, ExcelStyles)


class TestUtils:
    """Testy funkcji narzędziowych."""
    
    def test_sanitize_filename_simple(self):
        """Proste sanityzowanie nazwy pliku."""
        assert sanitize_filename("test.xlsx") == "test.xlsx"
    
    def test_sanitize_filename_removes_path_components(self):
        """Usuwa komponenty ścieżki."""
        assert sanitize_filename("../test.xlsx") == "test.xlsx"
    
    def test_validate_output_path_adds_extension(self):
        """Dodaje rozszerzenie .xlsx jeśli brakuje."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = validate_output_path("test", Path(tmpdir))
            assert path.suffix == ".xlsx"
    
    def test_validate_output_path_prevents_traversal(self):
        """Path traversal jest bezpiecznie obsługiwany."""
        with tempfile.TemporaryDirectory() as tmpdir:
            path = validate_output_path("../../etc/passwd", Path(tmpdir))
            assert str(path).startswith(tmpdir)
            assert path.name == "passwd.xlsx"


class TestSheets:
    """Testy tworzenia arkuszy."""
    
    def test_settings_sheet_creation(self):
        """Testuje tworzenie arkusza Ustawienia."""
        wb = Workbook()
        sheet = SettingsSheet(wb).create()
        
        assert sheet.title == "Ustawienia"
        # Sprawdź czy zawiera sekcję profilu mocy
        assert any("PROFIL MOCY" in str(cell.value or "") 
                   for row in sheet.iter_rows(max_row=20) for cell in row)
    
    def test_log_sheet_creation(self):
        """Testuje tworzenie arkusza Dziennik."""
        wb = Workbook()
        wb.active.title = "Temp"
        sheet = LogSheet(wb).create()
        
        assert sheet.title == "Dziennik"
        assert sheet.cell(row=1, column=1).value == "Data"
        # Sprawdź metryki WKO5
        headers = [sheet.cell(row=1, column=i).value for i in range(1, 43)]
        assert "TSS" in headers
        assert "CTL" in headers
    
    def test_dashboard_sheet_creation(self):
        """Testuje tworzenie arkusza Dashboard."""
        wb = Workbook()
        wb.active.title = "Temp"
        sheet = DashboardSheet(wb).create()
        
        assert sheet.title == "Dashboard"
        # Sprawdź PMC sekcję
        assert any("CTL" in str(cell.value or "") 
                   for row in sheet.iter_rows(max_row=20) for cell in row)
    
    def test_power_zones_sheet_creation(self):
        """Testuje tworzenie arkusza Strefy Mocy."""
        wb = Workbook()
        wb.active.title = "Temp"
        sheet = PowerZonesSheet(wb).create()
        
        assert sheet.title == "Strefy Mocy"
        # Sprawdź czy zawiera 7 stref
        zone_cells = [cell.value for row in sheet.iter_rows(max_row=15) 
                      for cell in row if cell.value and str(cell.value).startswith("Z")]
        assert len([z for z in zone_cells if z in ["Z1", "Z2", "Z3", "Z4", "Z5", "Z6", "Z7"]]) >= 7
    
    def test_cho_sources_sheet_creation(self):
        """Testuje tworzenie arkusza Źródła CHO."""
        wb = Workbook()
        wb.active.title = "Temp"
        sheet = CHOSourcesSheet(wb).create()
        
        assert sheet.title == "Źródła CHO"


class TestMain:
    """Testy głównego przepływu."""
    
    def test_create_workbook_returns_workbook(self):
        """Testuje czy create_workbook zwraca skoroszyt."""
        wb = create_workbook()
        assert isinstance(wb, Workbook)
    
    def test_create_workbook_has_all_sheets(self):
        """Testuje czy wszystkie arkusze zostały utworzone."""
        wb = create_workbook()
        sheet_names = wb.sheetnames
        
        assert "Ustawienia" in sheet_names
        assert "Dziennik" in sheet_names
        assert "Dashboard" in sheet_names
        assert "Strefy Mocy" in sheet_names
        assert "Źródła CHO" in sheet_names
    
    def test_create_workbook_sheet_count(self):
        """Testuje liczbę utworzonych arkuszy (5)."""
        wb = create_workbook()
        assert len(wb.sheetnames) == 5
    
    def test_full_workflow(self):
        """Testuje pełny przepływ: tworzenie i zapis."""
        wb = create_workbook()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            path = safe_save_workbook(wb, "full_test.xlsx", Path(tmpdir))
            
            assert path.exists()
            assert path.stat().st_size > 0
            
            # Sprawdź czy można otworzyć
            from openpyxl import load_workbook
            loaded = load_workbook(path)
            assert len(loaded.sheetnames) == 5


class TestFormulas:
    """Testy formuł Excel."""
    
    def test_log_sheet_has_pmc_formulas(self):
        """Sprawdza czy dziennik zawiera formuły PMC."""
        wb = Workbook()
        wb.active.title = "Temp"
        sheet = LogSheet(wb).create()
        
        # Sprawdź formuły TSS, CTL, ATL, TSB
        tss_cell = sheet['U2']
        assert tss_cell.value is not None and str(tss_cell.value).startswith('=')
        
        ctl_cell = sheet['X2']
        assert ctl_cell.value is not None and str(ctl_cell.value).startswith('=')
    
    def test_power_zones_has_ftp_formulas(self):
        """Sprawdza czy strefy mocy używają FTP."""
        wb = Workbook()
        wb.active.title = "Temp"
        sheet = PowerZonesSheet(wb).create()
        
        # Sprawdź czy kolumna Min W ma formułę
        min_w_cell = sheet.cell(row=6, column=5)  # Pierwsza strefa, kolumna Min W
        assert min_w_cell.value is not None and str(min_w_cell.value).startswith('=')


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
