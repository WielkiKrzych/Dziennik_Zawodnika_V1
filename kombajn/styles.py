"""
Style Excel dla Kombajnu Triathlonisty.

Ten moduł definiuje wszystkie style używane w arkuszach Excel,
zapewniając spójność wizualną i łatwość modyfikacji.
"""

from dataclasses import dataclass, field
from typing import Optional

from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from kombajn.config import COLORS


@dataclass
class ExcelStyles:
    """
    Kolekcja stylów Excel używanych w całym skoroszycie.
    
    Attributes:
        header_font: Czcionka dla nagłówków (pogrubiona, biała)
        header_fill: Wypełnienie nagłówków (niebieski)
        header_align: Wyrównanie nagłówków (środek, zawijanie)
        input_fill: Wypełnienie komórek do wpisywania (żółty)
        formula_fill: Wypełnienie komórek z formułami (szary)
        formula_font: Czcionka dla formuł (pogrubiona)
        info_font: Czcionka dla informacji (kursywa, szara)
        thin_border: Cienka ramka
        thick_right_border: Gruba prawa ramka (koniec sekcji)
    """
    
    # Nagłówki
    header_font: Font = field(
        default_factory=lambda: Font(bold=True, color=COLORS.HEADER_TEXT)
    )
    header_fill: PatternFill = field(
        default_factory=lambda: PatternFill(
            start_color=COLORS.HEADER_BG, 
            end_color=COLORS.HEADER_BG, 
            fill_type="solid"
        )
    )
    header_align: Alignment = field(
        default_factory=lambda: Alignment(
            horizontal="center", 
            vertical="center", 
            wrap_text=True
        )
    )
    
    # Komórki do wpisywania
    input_fill: PatternFill = field(
        default_factory=lambda: PatternFill(
            start_color=COLORS.INPUT_BG, 
            end_color=COLORS.INPUT_BG, 
            fill_type="solid"
        )
    )
    
    # Komórki z formułami
    formula_fill: PatternFill = field(
        default_factory=lambda: PatternFill(
            start_color=COLORS.FORMULA_BG, 
            end_color=COLORS.FORMULA_BG, 
            fill_type="solid"
        )
    )
    formula_font: Font = field(default_factory=lambda: Font(bold=True))
    
    # Tekst informacyjny
    info_font: Font = field(
        default_factory=lambda: Font(italic=True, color=COLORS.INFO_TEXT)
    )
    
    # Ramki
    thin_border: Border = field(
        default_factory=lambda: Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    )
    thick_right_border: Border = field(
        default_factory=lambda: Border(
            left=Side(style='thin'),
            right=Side(style='thick'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    )
    
    def apply_header_style(self, cell: Cell) -> None:
        """Aplikuje styl nagłówka do komórki."""
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_align
        cell.border = self.thin_border
    
    def apply_input_style(self, cell: Cell) -> None:
        """Aplikuje styl komórki do wpisywania."""
        cell.fill = self.input_fill
        cell.border = self.thin_border
    
    def apply_formula_style(self, cell: Cell) -> None:
        """Aplikuje styl komórki z formułą."""
        cell.font = self.formula_font
        cell.fill = self.formula_fill
        cell.border = self.thin_border
    
    def apply_info_style(self, cell: Cell) -> None:
        """Aplikuje styl tekstu informacyjnego."""
        cell.font = self.info_font
    
    def get_section_border(self, is_section_end: bool) -> Border:
        """
        Zwraca odpowiednią ramkę w zależności od pozycji w sekcji.
        
        Args:
            is_section_end: Czy komórka jest na końcu sekcji logicznej
            
        Returns:
            Gruba prawa ramka dla końca sekcji, cienka w pozostałych przypadkach
        """
        return self.thick_right_border if is_section_end else self.thin_border


# Singleton dla stylów - używany w całej aplikacji
DEFAULT_STYLES = ExcelStyles()
