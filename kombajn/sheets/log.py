"""
Arkusz Dziennik.

Rozszerzony dziennik kolarza z metrykami WKO5 (TSS, IF, NP) i PMC (CTL, ATL, TSB).
"""

from typing import Dict, List

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from kombajn.config import (
    LOG_HEADERS,
    LOG_INPUT_COLUMNS,
    LOG_SECTION_END_COLUMNS,
    LOG_COLUMN_WIDTHS,
    SHEET_CONFIG,
)
from kombajn.sheets.base import BaseSheet


class LogSheet(BaseSheet):
    """
    Arkusz dziennika kolarskiego z metrykami WKO5/INSCYD.
    
    Sekcje kolumn:
    - Ogólne (Data, Tydzień, Dzień)
    - Fizjologia (Waga, RHR, HRV, Sen)
    - Dane treningu (Czas, Dystans, Power, HR)
    - Metryki WKO5 (IF, TSS, W/kg, Strefa)
    - PMC (CTL, ATL, TSB)
    - Kalorie i makroskładniki
    - CHO podczas treningu
    - Notatki
    """
    
    def __init__(self, workbook: Workbook) -> None:
        """Inicjalizuje arkusz Dziennik."""
        super().__init__(workbook, "Dziennik")
    
    def create(self) -> Worksheet:
        """
        Tworzy arkusz Dziennik.
        
        Returns:
            Utworzony arkusz
        """
        ws = self._create_worksheet()
        
        self._add_headers(ws)
        self._format_data_rows(ws)
        self._add_formulas(ws)
        self._add_date_column(ws)
        self._set_column_widths(LOG_COLUMN_WIDTHS)
        
        # Zamrożenie pierwszego wiersza i pierwszych 3 kolumn
        ws.freeze_panes = 'D2'
        
        return ws
    
    def _add_headers(self, ws: Worksheet) -> None:
        """Dodaje nagłówki kolumn."""
        for i, header in enumerate(LOG_HEADERS, 1):
            cell = ws.cell(row=1, column=i)
            cell.value = header
            
            # Styl nagłówka
            self.styles.apply_header_style(cell)
            
            # Gruba ramka dla końca sekcji
            if i in LOG_SECTION_END_COLUMNS:
                cell.border = self.styles.thick_right_border
    
    def _format_data_rows(self, ws: Worksheet) -> None:
        """Formatuje wiersze danych (żółte/szare tło, ramki)."""
        max_rows = SHEET_CONFIG.MAX_LOG_ROWS + 1
        
        for i in range(1, len(LOG_HEADERS) + 1):
            is_input = i in LOG_INPUT_COLUMNS
            is_section_end = i in LOG_SECTION_END_COLUMNS
            border = self.styles.get_section_border(is_section_end)
            fill = self.styles.input_fill if is_input else self.styles.formula_fill
            
            for row in range(2, max_rows + 1):
                cell = ws.cell(row=row, column=i)
                cell.fill = fill
                cell.border = border
    
    def _add_formulas(self, ws: Worksheet) -> None:
        """Dodaje formuły do wiersza 2 i kopiuje w dół."""
        # Mapowanie kolumn (1-based):
        # 1=Data, 2=Tydzień, 3=Dzień tyg
        # 4=Waga, 5=Waga śr, 6=RHR, 7=HRV, 8=Sen, 9=Jakość snu, 10=Samopoczucie
        # 11=Czas, 12=Dystans, 13=Przewyższ, 14=AvgP, 15=NP, 16=MaxP, 17=Kadencja, 18=AvgHR, 19=MaxHR
        # 20=IF, 21=TSS, 22=W/kg, 23=Strefa
        # 24=CTL, 25=ATL, 26=TSB
        # 27=Kcal tren, 28=TDEE, 29=CEL, 30=Spożyte, 31=Bilans
        # 32=CEL B, 33=CEL T, 34=CEL W
        # 35=Spoż B, 36=Spoż T, 37=Spoż W
        # 38=CHO/h, 39=Nawodnienie
        # 40=Typ, 41=RPE, 42=Notatki
        
        formulas: Dict[str, str] = {
            # === SEKCJA OGÓLNE ===
            # Tydzień
            'B2': '=IF(ISNUMBER(A2), WEEKNUM(A2, 2), "")',
            # Dzień tygodnia
            'C2': '=IF(ISNUMBER(A2), TEXT(A2, "ddd"), "")',
            
            # === SEKCJA FIZJOLOGIA ===
            # Waga średnia 7-dniowa
            'E2': '=IF(ISNUMBER(D2), IFERROR(AVERAGE(INDIRECT("D"&MAX(2,ROW()-6)&":D"&ROW())), D2), "")',
            
            # === SEKCJA METRYKI WKO5 ===
            # IF = NP / FTP
            'T2': '=IF(AND(ISNUMBER(O2), Ustawienia!$B$6>0), O2/Ustawienia!$B$6, "")',
            
            # TSS = (czas_sek * NP * IF) / (FTP * 3600) * 100
            # czas jest w minutach (K2), konwersja: K2 * 60
            'U2': '=IF(AND(ISNUMBER(K2), ISNUMBER(O2), ISNUMBER(T2), Ustawienia!$B$6>0), '
                  '(K2*60*O2*T2)/(Ustawienia!$B$6*3600)*100, "")',
            
            # W/kg (NP)
            'V2': '=IF(AND(ISNUMBER(O2), Ustawienia!$B$3>0), O2/Ustawienia!$B$3, "")',
            
            # Strefa dominująca (wg NP i FTP)
            'W2': ('=IF(T2="", "", '
                   'IF(T2<0.55, "Z1", '
                   'IF(T2<0.75, "Z2", '
                   'IF(T2<0.90, "Z3", '
                   'IF(T2<1.05, "Z4", '
                   'IF(T2<1.20, "Z5", '
                   'IF(T2<1.50, "Z6", "Z7")))))))'),
            
            # === SEKCJA PMC ===
            # CTL (42-day average TSS)
            'X2': ('=IFERROR(AVERAGE(INDIRECT("U"&MAX(2,ROW()-41)&":U"&ROW())), "")'),
            
            # ATL (7-day average TSS)
            'Y2': ('=IFERROR(AVERAGE(INDIRECT("U"&MAX(2,ROW()-6)&":U"&ROW())), "")'),
            
            # TSB = CTL - ATL (forma: + = świeży, - = zmęczony)
            'Z2': '=IF(AND(ISNUMBER(X2), ISNUMBER(Y2)), X2-Y2, "")',
            
            # === SEKCJA KALORIE ===
            # Kcal treningu (szacunek z TSS lub manual)
            'AA2': '=IF(ISNUMBER(U2), ROUND(U2 * Ustawienia!$B$6 / 100 * 3.6, 0), "")',
            
            # TDEE = CPM + kcal treningu
            'AB2': '=IF(ISNUMBER(AA2), Ustawienia!$B$26 + AA2, Ustawienia!$B$26)',
            
            # CEL Kcal = TDEE - deficyt
            'AC2': '=AB2 - Ustawienia!$B$27',
            
            # Bilans = Spożyte - Cel
            'AE2': '=IF(ISBLANK(AD2), "", AD2 - AC2)',
            
            # === SEKCJA MAKRO ===
            # CEL Białko = współczynnik * waga
            'AF2': ('=IF(OR(Ustawienia!$B$3="", Ustawienia!$B$3=0), "", '
                   'ROUND(Ustawienia!$B$29 * Ustawienia!$B$3, 0))'),
            
            # CEL Tłuszcze = % TDEE / 9
            'AG2': '=IFERROR(ROUND((AC2 * Ustawienia!$B$30) / 9, 0), "")',
            
            # CEL Węgle = pozostałe kcal / 4
            'AH2': '=IFERROR(ROUND((AC2 - (AF2*4) - (AG2*9)) / 4, 0), "")',
        }
        
        # Dodaj formuły do wiersza 2
        for cell_ref, formula in formulas.items():
            ws[cell_ref] = formula
            ws[cell_ref].font = self.styles.formula_font
        
        # Ustaw formaty liczb
        ws['T2'].number_format = '0.00'   # IF
        ws['U2'].number_format = '0'      # TSS
        ws['V2'].number_format = '0.00'   # W/kg
        ws['X2'].number_format = '0'      # CTL
        ws['Y2'].number_format = '0'      # ATL
        ws['Z2'].number_format = '+0;-0;0'  # TSB z plusem
    
    def _add_date_column(self, ws: Worksheet) -> None:
        """Dodaje kolumnę dat z automatycznym wypełnianiem."""
        # A2 - data startowa (do wpisania)
        ws['A2'] = ""
        ws['A2'].fill = self.styles.input_fill
        ws['A2'].number_format = 'yyyy-mm-dd'
        
        # Automatyczne wypełnianie kolejnych dat
        for row in range(3, SHEET_CONFIG.INITIAL_DAYS_COUNT + 2):
            ws[f'A{row}'] = '=IF(ISBLANK($A$2), "", $A$2 + (ROW()-2))'
            ws[f'A{row}'].number_format = 'yyyy-mm-dd'
