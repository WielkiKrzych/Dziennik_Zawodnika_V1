"""
Bazowa klasa abstrakcyjna dla arkuszy Excel.

Definiuje wspólny interfejs dla wszystkich arkuszy w skoroszycie.
"""

from abc import ABC, abstractmethod
from typing import Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from kombajn.styles import ExcelStyles, DEFAULT_STYLES


class BaseSheet(ABC):
    """
    Abstrakcyjna klasa bazowa dla wszystkich arkuszy.
    
    Attributes:
        workbook: Skoroszyt Excel
        worksheet: Arkusz roboczy
        styles: Kolekcja stylów Excel
        title: Nazwa arkusza
    """
    
    def __init__(
        self,
        workbook: Workbook,
        title: str,
        styles: Optional[ExcelStyles] = None
    ) -> None:
        """
        Inicjalizuje arkusz.
        
        Args:
            workbook: Skoroszyt Excel
            title: Nazwa arkusza
            styles: Opcjonalna kolekcja stylów (domyślnie DEFAULT_STYLES)
        """
        self.workbook = workbook
        self.title = title
        self.styles = styles or DEFAULT_STYLES
        self.worksheet: Optional[Worksheet] = None
    
    @abstractmethod
    def create(self) -> Worksheet:
        """
        Tworzy i konfiguruje arkusz.
        
        Returns:
            Utworzony arkusz roboczy
        """
        pass
    
    def _create_worksheet(self, use_active: bool = False) -> Worksheet:
        """
        Tworzy nowy arkusz lub używa aktywnego.
        
        Args:
            use_active: Czy użyć aktywnego arkusza (dla pierwszego arkusza)
            
        Returns:
            Arkusz roboczy
        """
        if use_active:
            ws = self.workbook.active
            ws.title = self.title
        else:
            ws = self.workbook.create_sheet(self.title)
        
        self.worksheet = ws
        return ws
    
    def _set_column_widths(self, widths: list[int]) -> None:
        """
        Ustawia szerokości kolumn.
        
        Args:
            widths: Lista szerokości dla kolejnych kolumn
        """
        from openpyxl.utils import get_column_letter
        
        if self.worksheet is None:
            return
        
        for i, width in enumerate(widths, 1):
            self.worksheet.column_dimensions[get_column_letter(i)].width = width
