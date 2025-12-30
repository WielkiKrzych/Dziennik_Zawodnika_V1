"""
Arkusz Dashboard.

Rozszerzony dashboard z PMC (Performance Management Chart) i podsumowaniami kolarskimi.
"""

import datetime
from typing import Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from kombajn.config import COLORS
from kombajn.sheets.base import BaseSheet


class DashboardSheet(BaseSheet):
    """
    Arkusz dashboardu z PMC i podsumowaniami.
    
    Sekcje:
    - PMC (Performance Management Chart) - CTL, ATL, TSB
    - Podsumowanie tygodniowe (TSS, dystans, czas)
    - Wskaźniki trendu
    - Instrukcje
    """
    
    def __init__(self, workbook: Workbook) -> None:
        """Inicjalizuje arkusz Dashboard."""
        super().__init__(workbook, "Dashboard")
    
    def create(self) -> Worksheet:
        """
        Tworzy arkusz Dashboard.
        
        Returns:
            Utworzony arkusz
        """
        ws = self._create_worksheet()
        
        current_row = 1
        current_row = self._add_pmc_section(ws, current_row)
        current_row = self._add_weekly_summary(ws, current_row + 2)
        current_row = self._add_monthly_summary(ws, current_row + 2)
        self._add_instructions(ws, current_row + 2)
        
        self._set_column_widths([25, 15, 15, 15, 40])
        
        return ws
    
    def _add_section_header(self, ws: Worksheet, row: int, title: str, 
                           cols: int = 5) -> None:
        """Dodaje nagłówek sekcji."""
        cell = ws.cell(row=row, column=1)
        cell.value = title
        cell.font = Font(bold=True, size=14, color=COLORS.HEADER_TEXT)
        cell.fill = PatternFill(start_color=COLORS.HEADER_BG, 
                                 end_color=COLORS.HEADER_BG, fill_type="solid")
        ws.merge_cells(f'A{row}:{chr(64+cols)}{row}')
        ws.row_dimensions[row].height = 28
    
    def _add_pmc_section(self, ws: Worksheet, start_row: int) -> int:
        """Dodaje sekcję PMC (Performance Management Chart)."""
        self._add_section_header(ws, start_row, "📈 PERFORMANCE MANAGEMENT CHART (PMC)")
        
        row = start_row + 2
        
        # Nagłówki PMC
        headers = ["Metryka", "Wartość", "Trend", "Status", "Opis"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            self.styles.apply_header_style(cell)
        
        row += 1
        
        # CTL (Fitness)
        ws.cell(row=row, column=1).value = "CTL (Fitness)"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = (
            "=IFERROR(INDEX('Dziennik'!X:X, MATCH(9.99E+307, 'Dziennik'!X:X)), \"--\")"
        )
        ws.cell(row=row, column=2).fill = PatternFill(
            start_color=COLORS.CTL_COLOR, end_color=COLORS.CTL_COLOR, fill_type="solid"
        )
        ws.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=2).number_format = '0'
        ws.cell(row=row, column=5).value = "Średni TSS z 42 dni - wskaźnik kondycji"
        self.styles.apply_info_style(ws.cell(row=row, column=5))
        
        row += 1
        
        # ATL (Fatigue)
        ws.cell(row=row, column=1).value = "ATL (Fatigue)"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = (
            "=IFERROR(INDEX('Dziennik'!Y:Y, MATCH(9.99E+307, 'Dziennik'!Y:Y)), \"--\")"
        )
        ws.cell(row=row, column=2).fill = PatternFill(
            start_color=COLORS.ATL_COLOR, end_color=COLORS.ATL_COLOR, fill_type="solid"
        )
        ws.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=2).number_format = '0'
        ws.cell(row=row, column=5).value = "Średni TSS z 7 dni - wskaźnik zmęczenia"
        self.styles.apply_info_style(ws.cell(row=row, column=5))
        
        row += 1
        
        # TSB (Form)
        ws.cell(row=row, column=1).value = "TSB (Form)"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = (
            "=IFERROR(INDEX('Dziennik'!Z:Z, MATCH(9.99E+307, 'Dziennik'!Z:Z)), \"--\")"
        )
        ws.cell(row=row, column=2).fill = PatternFill(
            start_color=COLORS.TSB_COLOR, end_color=COLORS.TSB_COLOR, fill_type="solid"
        )
        ws.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=2).number_format = '+0;-0;0'
        # Status formy
        ws.cell(row=row, column=4).value = (
            '=IF(B' + str(row) + '="--", "", '
            'IF(B' + str(row) + '>25, "⚠️ Przetrenowanie?", '
            'IF(B' + str(row) + '>10, "🟢 Świeży", '
            'IF(B' + str(row) + '>-10, "🟡 Neutralny", '
            'IF(B' + str(row) + '>-25, "🟠 Zmęczony", "🔴 Bardzo zmęczony")))))'
        )
        ws.cell(row=row, column=5).value = "CTL - ATL: + = świeży, - = zmęczony"
        self.styles.apply_info_style(ws.cell(row=row, column=5))
        
        row += 2
        
        # Legenda TSB
        ws.cell(row=row, column=1).value = "📊 Interpretacja TSB:"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        tsb_legend = [
            ("> +25", "Zbyt wypoczęty - tracisz formę", "⚠️"),
            ("+10 do +25", "Optymalny na wyścig/test", "🟢"),
            ("-10 do +10", "Produktywny trening", "🟡"),
            ("-25 do -10", "Ciężki blok treningowy", "🟠"),
            ("< -25", "Ryzyko przetrenowania!", "🔴"),
        ]
        
        for tsb_range, desc, icon in tsb_legend:
            ws.cell(row=row, column=1).value = f"{icon} TSB {tsb_range}"
            ws.cell(row=row, column=2).value = desc
            ws.merge_cells(f'B{row}:E{row}')
            self.styles.apply_info_style(ws.cell(row=row, column=2))
            row += 1
        
        return row
    
    def _add_weekly_summary(self, ws: Worksheet, start_row: int) -> int:
        """Dodaje sekcję podsumowania tygodniowego."""
        self._add_section_header(ws, start_row, "📅 PODSUMOWANIE TYGODNIOWE")
        
        row = start_row + 2
        
        # Wybór tygodnia
        ws.cell(row=row, column=1).value = "Wybierz tydzień:"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = datetime.date.today().isocalendar()[1]
        ws.cell(row=row, column=2).fill = self.styles.input_fill
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        
        row += 2
        
        # Nagłówki
        headers = ["Metryka", "Wartość", "Jednostka"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            self.styles.apply_header_style(cell)
        
        row += 1
        
        # Dane tygodniowe
        weekly_metrics = [
            ("Suma TSS", "=IFERROR(SUMIFS('Dziennik'!U:U, 'Dziennik'!B:B, $B$" + 
             str(start_row + 2) + "), \"--\")", "TSS"),
            ("Suma czasu jazdy", "=IFERROR(SUMIFS('Dziennik'!K:K, 'Dziennik'!B:B, $B$" + 
             str(start_row + 2) + ")/60, \"--\")", "h"),
            ("Suma dystansu", "=IFERROR(SUMIFS('Dziennik'!L:L, 'Dziennik'!B:B, $B$" + 
             str(start_row + 2) + "), \"--\")", "km"),
            ("Suma przewyższeń", "=IFERROR(SUMIFS('Dziennik'!M:M, 'Dziennik'!B:B, $B$" + 
             str(start_row + 2) + "), \"--\")", "m"),
            ("Średni IF", "=IFERROR(AVERAGEIFS('Dziennik'!T:T, 'Dziennik'!B:B, $B$" + 
             str(start_row + 2) + "), \"--\")", ""),
            ("Średnia NP", "=IFERROR(AVERAGEIFS('Dziennik'!O:O, 'Dziennik'!B:B, $B$" + 
             str(start_row + 2) + "), \"--\")", "W"),
            ("Średnia waga", "=IFERROR(AVERAGEIFS('Dziennik'!D:D, 'Dziennik'!B:B, $B$" + 
             str(start_row + 2) + "), \"--\")", "kg"),
            ("Liczba treningów", "=COUNTIFS('Dziennik'!B:B, $B$" + str(start_row + 2) + 
             ", 'Dziennik'!K:K, \">0\")", ""),
        ]
        
        for label, formula, unit in weekly_metrics:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).value = formula
            self.styles.apply_formula_style(ws.cell(row=row, column=2))
            if "IF" in label or "waga" in label.lower():
                ws.cell(row=row, column=2).number_format = '0.00'
            else:
                ws.cell(row=row, column=2).number_format = '0'
            ws.cell(row=row, column=3).value = unit
            row += 1
        
        return row
    
    def _add_monthly_summary(self, ws: Worksheet, start_row: int) -> int:
        """Dodaje sekcję podsumowania miesięcznego."""
        self._add_section_header(ws, start_row, "📆 STATYSTYKI CAŁKOWITE", cols=3)
        
        row = start_row + 2
        
        stats = [
            ("Suma TSS (wszystkie)", "=IFERROR(SUM('Dziennik'!U:U), 0)", "TSS"),
            ("Suma dystansu (wszystkie)", "=IFERROR(SUM('Dziennik'!L:L), 0)", "km"),
            ("Suma przewyższeń (wszystkie)", "=IFERROR(SUM('Dziennik'!M:M), 0)", "m"),
            ("Suma czasu (wszystkie)", "=IFERROR(SUM('Dziennik'!K:K)/60, 0)", "h"),
            ("Liczba dni treningowych", "=COUNTIF('Dziennik'!K:K, \">0\")", "dni"),
        ]
        
        for label, formula, unit in stats:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).value = formula
            self.styles.apply_formula_style(ws.cell(row=row, column=2))
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3).value = unit
            row += 1
        
        return row
    
    def _add_instructions(self, ws: Worksheet, start_row: int) -> None:
        """Dodaje instrukcje."""
        self._add_section_header(ws, start_row, "📖 INSTRUKCJA", cols=5)
        
        row = start_row + 2
        
        instructions = [
            "1. Wypełniaj dziennik codziennie - dane z Garmina/Zwifta",
            "2. Śledź TSB: -10 do +10 = produktywny trening, >+15 = gotowy na wyścig",
            "3. Tygodniowy TSS: amator 300-500, zaawansowany 500-800, pro 800-1200+",
            "4. Dla wykresów: zaznacz kolumny i wybierz Wstawianie → Wykres",
            "5. FTP aktualizuj co 4-6 tyg lub po teście",
        ]
        
        for instruction in instructions:
            ws.cell(row=row, column=1).value = instruction
            self.styles.apply_info_style(ws.cell(row=row, column=1))
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
