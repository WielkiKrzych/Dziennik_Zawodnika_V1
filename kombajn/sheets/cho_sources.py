"""
Arkusz Źródła CHO.

Baza produktów węglowodanowych dla kolarzy z rozszerzonymi danymi.
"""

from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from kombajn.config import (
    CHO_HEADERS,
    CHO_COLUMN_WIDTHS,
    CHO_SAMPLE_DATA,
    SHEET_CONFIG,
    COLORS,
)
from kombajn.sheets.base import BaseSheet


class CHOSourcesSheet(BaseSheet):
    """
    Arkusz źródeł węglowodanów dla kolarzy.
    
    Rozszerzony o:
    - Szybkość wchłaniania
    - Produkty typowe dla kolarstwa (żele, napoje, mix MD:FR)
    - Planowanie CHO/h na trening
    """
    
    def __init__(self, workbook: Workbook) -> None:
        """Inicjalizuje arkusz Źródła CHO."""
        super().__init__(workbook, "Źródła CHO")
    
    def create(self) -> Worksheet:
        """
        Tworzy arkusz Źródła CHO.
        
        Returns:
            Utworzony arkusz
        """
        ws = self._create_worksheet()
        
        self._add_title(ws)
        self._add_headers(ws)
        self._format_input_columns(ws)
        self._add_sample_data(ws)
        self._add_cho_calculator(ws)
        self._set_column_widths(CHO_COLUMN_WIDTHS)
        
        ws.freeze_panes = 'A3'
        
        return ws
    
    def _add_title(self, ws: Worksheet) -> None:
        """Dodaje tytuł arkusza."""
        ws['A1'] = "🍌 ŹRÓDŁA WĘGLOWODANÓW - BAZA DLA KOLARZA"
        ws['A1'].font = Font(bold=True, size=14, color=COLORS.HEADER_TEXT)
        ws['A1'].fill = PatternFill(start_color=COLORS.HEADER_BG, 
                                     end_color=COLORS.HEADER_BG, fill_type="solid")
        ws.merge_cells('A1:I1')
        ws.row_dimensions[1].height = 28
    
    def _add_headers(self, ws: Worksheet) -> None:
        """Dodaje nagłówki kolumn."""
        for i, header in enumerate(CHO_HEADERS, 1):
            cell = ws.cell(row=2, column=i)
            cell.value = header
            self.styles.apply_header_style(cell)
    
    def _format_input_columns(self, ws: Worksheet) -> None:
        """Formatuje kolumny do wpisywania (żółte tło)."""
        input_cols = [1, 2, 3, 4, 7, 8, 9]  # Nazwa, Porcja, CHO, kcal, Typ, Wchłanianie, Uwagi
        max_rows = 100
        
        for row in range(3, max_rows + 1):
            for col in input_cols:
                ws.cell(row=row, column=col).fill = self.styles.input_fill
    
    def _add_sample_data(self, ws: Worksheet) -> None:
        """Dodaje przykładowe dane produktów."""
        for row_idx, item in enumerate(CHO_SAMPLE_DATA, start=3):
            name, portion, cho_100g, kcal_100g, product_type, absorption, notes = item
            
            ws.cell(row=row_idx, column=1).value = name
            ws.cell(row=row_idx, column=2).value = portion
            ws.cell(row=row_idx, column=3).value = cho_100g
            ws.cell(row=row_idx, column=4).value = kcal_100g
            ws.cell(row=row_idx, column=7).value = product_type
            ws.cell(row=row_idx, column=8).value = absorption
            ws.cell(row=row_idx, column=9).value = notes
            
            # Formuły dla obliczonych kolumn
            self._add_portion_formulas(ws, row_idx)
    
    def _add_portion_formulas(self, ws: Worksheet, row: int) -> None:
        """Dodaje formuły obliczające wartości dla porcji."""
        # CHO w porcji
        cho_formula = f'=IF(OR(B{row}="",C{row}=""), "", ROUND(B{row} * C{row} / 100, 1))'
        ws.cell(row=row, column=5).value = cho_formula
        self.styles.apply_formula_style(ws.cell(row=row, column=5))
        
        # kcal w porcji
        kcal_formula = f'=IF(OR(B{row}="",D{row}=""), "", ROUND(B{row} * D{row} / 100, 0))'
        ws.cell(row=row, column=6).value = kcal_formula
        self.styles.apply_formula_style(ws.cell(row=row, column=6))
    
    def _add_cho_calculator(self, ws: Worksheet) -> None:
        """Dodaje kalkulator CHO na godzinę."""
        start_row = len(CHO_SAMPLE_DATA) + 5
        
        ws.cell(row=start_row, column=1).value = "🧮 KALKULATOR CHO NA TRENING"
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(f'A{start_row}:E{start_row}')
        
        row = start_row + 2
        
        # Cel CHO/h
        ws.cell(row=row, column=1).value = "Cel CHO/h (g):"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = 60
        ws.cell(row=row, column=2).fill = self.styles.input_fill
        
        row += 1
        ws.cell(row=row, column=1).value = "Czas treningu (h):"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = 3
        ws.cell(row=row, column=2).fill = self.styles.input_fill
        
        row += 1
        ws.cell(row=row, column=1).value = "Całkowite CHO potrzebne:"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = f"=B{start_row+2}*B{start_row+3}"
        self.styles.apply_formula_style(ws.cell(row=row, column=2))
        ws.cell(row=row, column=3).value = "g"
        
        row += 2
        
        # Wskazówki
        tips = [
            "💡 Wskazówki spożywania CHO podczas jazdy:",
            "• Z2/Z3 < 2h: 30-40g/h (lub bez)",
            "• Z3/Z4 2-3h: 60g/h",
            "• Wyścig/intensywny: 80-90g/h (mix glukoza:fruktoza 1:0.8)",
            "• Ultra >5h: do 120g/h (wymaga treningu jelit!)",
            "• Zacznij od 30g/h i zwiększaj o 10g/h co tydzień",
        ]
        
        for tip in tips:
            ws.cell(row=row, column=1).value = tip
            ws.merge_cells(f'A{row}:E{row}')
            if tip.startswith("💡"):
                ws.cell(row=row, column=1).font = Font(bold=True)
            else:
                self.styles.apply_info_style(ws.cell(row=row, column=1))
            row += 1
