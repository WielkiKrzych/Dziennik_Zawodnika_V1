"""
Arkusz Ustawienia.

Zawiera profil użytkownika, parametry mocy (WKO5) i profil metaboliczny (INSCYD).
"""

from typing import Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from kombajn.config import (
    DEFAULTS, POWER_DEFAULTS, METABOLIC_DEFAULTS, COLORS
)
from kombajn.sheets.base import BaseSheet


class SettingsSheet(BaseSheet):
    """
    Arkusz z ustawieniami i profilem użytkownika.
    
    Sekcje:
    - Profil użytkownika (waga, wzrost)
    - Profil mocy WKO5 (FTP, HR)
    - Profil metaboliczny INSCYD (VO2max, VLaMax)
    - Cele kaloryczne i makroskładnikowe
    """
    
    def __init__(self, workbook: Workbook) -> None:
        """Inicjalizuje arkusz Ustawienia."""
        super().__init__(workbook, "Ustawienia")
    
    def create(self) -> Worksheet:
        """
        Tworzy arkusz Ustawienia.
        
        Returns:
            Utworzony arkusz
        """
        ws = self._create_worksheet(use_active=True)
        
        current_row = 1
        current_row = self._add_user_profile(ws, current_row)
        current_row = self._add_power_profile(ws, current_row + 1)
        current_row = self._add_metabolic_profile(ws, current_row + 1)
        current_row = self._add_calorie_settings(ws, current_row + 1)
        current_row = self._add_macro_targets(ws, current_row + 1)
        
        self._set_column_widths([35, 15, 35])
        
        return ws
    
    def _add_section_header(self, ws: Worksheet, row: int, title: str) -> None:
        """Dodaje nagłówek sekcji."""
        cell = ws.cell(row=row, column=1)
        cell.value = title
        cell.font = Font(bold=True, size=13, color=COLORS.HEADER_TEXT)
        cell.fill = PatternFill(start_color=COLORS.HEADER_BG, 
                                 end_color=COLORS.HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(vertical="center")
        ws.merge_cells(f'A{row}:C{row}')
        ws.row_dimensions[row].height = 25
    
    def _add_input_row(self, ws: Worksheet, row: int, label: str, 
                       value, info: str = "", number_format: str = None) -> None:
        """Dodaje wiersz z polem do wpisania."""
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        cell = ws.cell(row=row, column=2)
        cell.value = value
        cell.fill = self.styles.input_fill
        if number_format:
            cell.number_format = number_format
        
        if info:
            ws.cell(row=row, column=3).value = info
            self.styles.apply_info_style(ws.cell(row=row, column=3))
    
    def _add_formula_row(self, ws: Worksheet, row: int, label: str, 
                         formula: str, info: str = "", number_format: str = None) -> None:
        """Dodaje wiersz z formułą."""
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        cell = ws.cell(row=row, column=2)
        cell.value = formula
        self.styles.apply_formula_style(cell)
        if number_format:
            cell.number_format = number_format
        
        if info:
            ws.cell(row=row, column=3).value = info
            self.styles.apply_info_style(ws.cell(row=row, column=3))
    
    def _add_user_profile(self, ws: Worksheet, start_row: int) -> int:
        """Dodaje sekcję profilu użytkownika."""
        self._add_section_header(ws, start_row, "📊 PROFIL UŻYTKOWNIKA")
        
        row = start_row + 1
        self._add_input_row(ws, row, "Imię / Pseudonim", "", "")
        row += 1
        self._add_input_row(ws, row, "Waga (kg)", POWER_DEFAULTS.WEIGHT_KG, 
                           "Aktualna waga do obliczeń W/kg", "0.0")
        row += 1
        self._add_input_row(ws, row, "Wzrost (cm)", 175, "")
        
        return row
    
    def _add_power_profile(self, ws: Worksheet, start_row: int) -> int:
        """Dodaje sekcję profilu mocy (WKO5)."""
        self._add_section_header(ws, start_row, "⚡ PROFIL MOCY (WKO5)")
        
        row = start_row + 1
        self._add_input_row(ws, row, "FTP (W)", POWER_DEFAULTS.FTP, 
                           "Functional Threshold Power")
        row += 1
        self._add_formula_row(ws, row, "W/kg", "=IF(B3>0, B6/B3, \"\")", 
                             "Automatycznie z FTP / waga", "0.00")
        row += 1
        self._add_input_row(ws, row, "HR Max (bpm)", POWER_DEFAULTS.MAX_HR, 
                           "Tętno maksymalne")
        row += 1
        self._add_input_row(ws, row, "HR Rest (bpm)", POWER_DEFAULTS.RESTING_HR, 
                           "Tętno spoczynkowe")
        row += 1
        self._add_input_row(ws, row, "Max Power 5s (W)", "", 
                           "Opcjonalnie - do profilu mocy")
        row += 1
        self._add_input_row(ws, row, "Max Power 1min (W)", "", 
                           "Opcjonalnie - do profilu mocy")
        row += 1
        self._add_input_row(ws, row, "Max Power 5min (W)", "", 
                           "Opcjonalnie - VO2max power")
        row += 1
        self._add_input_row(ws, row, "Max Power 20min (W)", "", 
                           "FTP ≈ 95% tej wartości")
        
        return row
    
    def _add_metabolic_profile(self, ws: Worksheet, start_row: int) -> int:
        """Dodaje sekcję profilu metabolicznego (INSCYD)."""
        self._add_section_header(ws, start_row, "🔬 PROFIL METABOLICZNY (INSCYD)")
        
        row = start_row + 1
        self._add_input_row(ws, row, "VO2max (ml/kg/min)", METABOLIC_DEFAULTS.VO2MAX, 
                           "Z testu INSCYD lub szacunkowo", "0.0")
        row += 1
        self._add_input_row(ws, row, "VLaMax (mmol/L/s)", METABOLIC_DEFAULTS.VLAMAX, 
                           "Maks. produkcja mleczanu", "0.00")
        row += 1
        self._add_formula_row(ws, row, "FatMax (W)", 
                             f"=ROUND(B6*{METABOLIC_DEFAULTS.FATMAX_PERCENT}, 0)", 
                             "Moc przy max spalaniu tłuszczu (~55% FTP)")
        row += 1
        self._add_formula_row(ws, row, "FatMax Zone", 
                             "=CONCATENATE(ROUND(B6*0.50,0), \" - \", ROUND(B6*0.65,0), \" W\")", 
                             "Strefa max spalania tłuszczu")
        row += 1
        self._add_input_row(ws, row, "Próg mleczanowy (W)", "", 
                           "LT1 / VT1 jeśli znany")
        row += 1
        self._add_input_row(ws, row, "Próg anaerobowy (W)", "", 
                           "LT2 / VT2 / MLSS jeśli znany")
        
        return row
    
    def _add_calorie_settings(self, ws: Worksheet, start_row: int) -> int:
        """Dodaje sekcję ustawień kalorycznych."""
        self._add_section_header(ws, start_row, "🔥 USTAWIENIA KALORYCZNE")
        
        row = start_row + 1
        self._add_input_row(ws, row, "BMR (kcal)", DEFAULTS.BMR, 
                           "Basal Metabolic Rate")
        row += 1
        self._add_input_row(ws, row, "TEF (kcal)", DEFAULTS.TEF, 
                           "Thermic Effect of Food (~10% BMR)")
        row += 1
        self._add_input_row(ws, row, "NEAT (kcal)", DEFAULTS.NEAT, 
                           "Non-Exercise Activity")
        row += 1
        self._add_formula_row(ws, row, "CPM (Baza)", 
                             f"=SUM(B{row-3}:B{row-1})", 
                             "Całkowita Przemiana Materii bez treningu")
        row += 1
        self._add_input_row(ws, row, "Cel (deficyt/nadwyżka)", DEFAULTS.DEFICIT, 
                           "Ujemna = deficyt, dodatnia = nadwyżka")
        
        return row
    
    def _add_macro_targets(self, ws: Worksheet, start_row: int) -> int:
        """Dodaje sekcję celów makroskładnikowych."""
        self._add_section_header(ws, start_row, "🥗 CELE MAKROSKŁADNIKOWE")
        
        row = start_row + 1
        self._add_input_row(ws, row, "Białko (g / kg mc)", DEFAULTS.PROTEIN_RATIO, 
                           "1.6-2.2g dla sportowców", "0.0")
        row += 1
        self._add_input_row(ws, row, "Tłuszcze (% TDEE)", DEFAULTS.FAT_RATIO, 
                           "0.20-0.30 (20-30%)", "0%")
        row += 1
        self._add_input_row(ws, row, "CHO podczas treningu (g/h)", 60, 
                           "60-90g/h dla intensywnych jazd")
        row += 1
        
        # Info o celach CHO
        ws.cell(row=row, column=1).value = "💡 Wskazówki CHO/h:"
        ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        tips = [
            "• Z2 (< 2h): 30-40g/h",
            "• Z3/Z4 (2-3h): 60g/h",
            "• Wyścig/intensywny: 80-90g/h",
            "• Ultra: do 120g/h (wymaga treningu jelit)",
        ]
        
        for tip in tips:
            ws.cell(row=row, column=1).value = tip
            self.styles.apply_info_style(ws.cell(row=row, column=1))
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
        
        return row
