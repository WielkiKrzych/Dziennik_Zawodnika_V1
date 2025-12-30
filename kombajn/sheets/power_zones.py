"""
Arkusz Strefy Mocy.

Tabela 7 stref mocy wg Coggan z automatycznym przeliczaniem z FTP.
"""

from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from kombajn.config import POWER_ZONES, COLORS
from kombajn.sheets.base import BaseSheet


class PowerZonesSheet(BaseSheet):
    """
    Arkusz stref mocy wg metodologii Coggan/WKO5.
    
    Zawiera:
    - 7 stref mocy z zakresami % FTP
    - Automatyczne przeliczanie na W z FTP
    - Opisy fizjologiczne każdej strefy
    - Strefy tętna
    """
    
    ZONE_COLORS = [
        COLORS.ZONE_1, COLORS.ZONE_2, COLORS.ZONE_3, COLORS.ZONE_4,
        COLORS.ZONE_5, COLORS.ZONE_6, COLORS.ZONE_7
    ]
    
    def __init__(self, workbook: Workbook) -> None:
        """Inicjalizuje arkusz Strefy Mocy."""
        super().__init__(workbook, "Strefy Mocy")
    
    def create(self) -> Worksheet:
        """
        Tworzy arkusz Strefy Mocy.
        
        Returns:
            Utworzony arkusz
        """
        ws = self._create_worksheet()
        
        self._add_title(ws)
        self._add_ftp_input(ws)
        self._add_power_zones_table(ws)
        self._add_hr_zones_table(ws)
        self._add_usage_notes(ws)
        self._set_column_widths([5, 25, 10, 10, 12, 12, 50])
        
        return ws
    
    def _add_title(self, ws: Worksheet) -> None:
        """Dodaje tytuł arkusza."""
        ws['A1'] = "STREFY MOCY (COGGAN / WKO5)"
        ws['A1'].font = Font(bold=True, size=16, color=COLORS.HEADER_TEXT)
        ws['A1'].fill = PatternFill(start_color=COLORS.HEADER_BG, 
                                     end_color=COLORS.HEADER_BG, fill_type="solid")
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 30
    
    def _add_ftp_input(self, ws: Worksheet) -> None:
        """Dodaje pole FTP."""
        ws['A3'] = "Twoje FTP (W):"
        ws['A3'].font = Font(bold=True, size=12)
        ws.merge_cells('A3:B3')
        
        ws['C3'] = "='Ustawienia'!$B$5"  # Pobiera FTP z ustawień
        ws['C3'].font = Font(bold=True, size=14)
        ws['C3'].fill = self.styles.formula_fill
        
        ws['D3'] = "W/kg:"
        ws['D3'].font = Font(bold=True)
        
        # W/kg = FTP / waga
        ws['E3'] = "=IF('Ustawienia'!$B$4>0, C3/'Ustawienia'!$B$4, \"\")"
        ws['E3'].font = Font(bold=True, size=12)
        ws['E3'].fill = self.styles.formula_fill
        ws['E3'].number_format = '0.00'
    
    def _add_power_zones_table(self, ws: Worksheet) -> None:
        """Dodaje tabelę stref mocy."""
        # Nagłówki tabeli
        headers = ["Strefa", "Nazwa", "Min %", "Max %", "Min W", "Max W", "Opis"]
        start_row = 5
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            self.styles.apply_header_style(cell)
        
        # Dane stref
        for i, zone in enumerate(POWER_ZONES):
            row = start_row + 1 + i
            zone_color = self.ZONE_COLORS[i]
            zone_fill = PatternFill(start_color=zone_color, end_color=zone_color, 
                                    fill_type="solid")
            
            # Strefa
            cell = ws.cell(row=row, column=1)
            cell.value = f"Z{zone.number}"
            cell.font = Font(bold=True, size=11)
            cell.fill = zone_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Nazwa
            cell = ws.cell(row=row, column=2)
            cell.value = zone.name
            cell.font = Font(bold=True)
            cell.fill = zone_fill
            cell.border = thin_border
            
            # Min %
            cell = ws.cell(row=row, column=3)
            cell.value = zone.min_pct
            cell.number_format = '0%'
            cell.fill = zone_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Max %
            cell = ws.cell(row=row, column=4)
            cell.value = zone.max_pct
            cell.number_format = '0%'
            cell.fill = zone_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Min W (formuła)
            cell = ws.cell(row=row, column=5)
            cell.value = f"=ROUND($C$3*C{row}, 0)"
            cell.font = Font(bold=True)
            cell.fill = zone_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Max W (formuła)
            cell = ws.cell(row=row, column=6)
            cell.value = f"=ROUND($C$3*D{row}, 0)"
            cell.font = Font(bold=True)
            cell.fill = zone_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            
            # Opis
            cell = ws.cell(row=row, column=7)
            cell.value = zone.description
            cell.fill = zone_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
        
        # Wysokość wierszy
        for i in range(len(POWER_ZONES)):
            ws.row_dimensions[start_row + 1 + i].height = 25
    
    def _add_hr_zones_table(self, ws: Worksheet) -> None:
        """Dodaje tabelę stref tętna."""
        start_row = 15
        
        ws.cell(row=start_row, column=1).value = "STREFY TĘTNA"
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(f'A{start_row}:G{start_row}')
        
        # HR Max input
        ws.cell(row=start_row + 1, column=1).value = "HR Max:"
        ws.cell(row=start_row + 1, column=1).font = Font(bold=True)
        ws.cell(row=start_row + 1, column=2).value = "='Ustawienia'!$B$7"
        ws.cell(row=start_row + 1, column=2).fill = self.styles.formula_fill
        
        # Nagłówki
        headers = ["Strefa", "Nazwa", "Min %", "Max %", "Min BPM", "Max BPM"]
        header_row = start_row + 3
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            self.styles.apply_header_style(cell)
        
        # Strefy tętna (5 stref)
        hr_zones = [
            (1, "Z1 - Regeneracja", 0.50, 0.60),
            (2, "Z2 - Wytrzymałość", 0.60, 0.70),
            (3, "Z3 - Tempo", 0.70, 0.80),
            (4, "Z4 - Próg", 0.80, 0.90),
            (5, "Z5 - VO2max", 0.90, 1.00),
        ]
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for i, (num, name, min_pct, max_pct) in enumerate(hr_zones):
            row = header_row + 1 + i
            
            ws.cell(row=row, column=1).value = f"Z{num}"
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            
            ws.cell(row=row, column=2).value = name
            ws.cell(row=row, column=2).border = thin_border
            
            ws.cell(row=row, column=3).value = min_pct
            ws.cell(row=row, column=3).number_format = '0%'
            ws.cell(row=row, column=3).border = thin_border
            
            ws.cell(row=row, column=4).value = max_pct
            ws.cell(row=row, column=4).number_format = '0%'
            ws.cell(row=row, column=4).border = thin_border
            
            hr_cell_ref = f"$B${start_row + 1}"
            ws.cell(row=row, column=5).value = f"=ROUND({hr_cell_ref}*C{row}, 0)"
            ws.cell(row=row, column=5).border = thin_border
            
            ws.cell(row=row, column=6).value = f"=ROUND({hr_cell_ref}*D{row}, 0)"
            ws.cell(row=row, column=6).border = thin_border
    
    def _add_usage_notes(self, ws: Worksheet) -> None:
        """Dodaje notatki z instrukcją."""
        start_row = 25
        
        ws.cell(row=start_row, column=1).value = "INSTRUKCJA"
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        
        notes = [
            "1. Ustaw swoje FTP w arkuszu [Ustawienia] - strefy przeliczą się automatycznie",
            "2. Z2-Z3: Większość treningu bazowego (70-80% czasu)",
            "3. Z4 Sweet Spot: Najbardziej efektywny trening dla FTP (88-94% FTP)",
            "4. Z5 VO2max: Interwały 3-8 min, rozwój wydolności tlenowej",
            "5. Monitoruj TSS: 300-500/tydzień dla amatorów, 700-1000+ dla zawodowców",
        ]
        
        for i, note in enumerate(notes):
            cell = ws.cell(row=start_row + 1 + i, column=1)
            cell.value = note
            cell.font = Font(italic=True, color=COLORS.INFO_TEXT)
            ws.merge_cells(f'A{start_row + 1 + i}:G{start_row + 1 + i}')
