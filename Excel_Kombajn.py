import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

print("Cześć. Zaczynam tworzyć Twój 'kombajn v2'...")

# --- Definicje Stylów ---
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
formula_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
formula_font = Font(bold=True)
info_font = Font(italic=True, color="808080")
thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))
thick_right_border = Border(left=Side(style='thin'),
                          right=Side(style='thick'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

# --- Tworzenie Skoroszytu ---
try:
    wb = openpyxl.Workbook()

    # --- 1. Zakładka: Ustawienia i Cele (bez zmian) ---
    ws_settings = wb.active
    ws_settings.title = "Ustawienia i Cele"
    print("Tworzę zakładkę [Ustawienia i Cele]...")

    settings_data = {
        'A1': "TWOJE STAŁE DANE", 'A2': "BMR (kcal)", 'A3': "TEF (kcal)",
        'A4': "NEAT (kcal)", 'A6': "CPM (Baza)", 'A8': "TWOJE CELE",
        'A9': "Planowany Deficyt (np. 500)", 'A11': "CEL: Białko (g / kgmc)",
        'A12': "CEL: Tłuszcze (% TDEE)", 'A14': "DANE ŚCIĄGANE Z DZIENNIKA",
        'A15': "Aktualna waga (ostatni wpis)",
    }

    for cell, value in settings_data.items():
        ws_settings[cell] = value
        if cell in ['A1', 'A8', 'A14']:
            ws_settings[cell].font = Font(bold=True, size=14)
            ws_settings.merge_cells(f'{cell}:{get_column_letter(3)}{cell[1:]}')
        else:
            ws_settings[cell].font = Font(bold=True)

    ws_settings['B6'] = "=SUM(B2:B4)"
    ws_settings['B6'].font = formula_font
    ws_settings['B6'].fill = formula_fill
    ws_settings['B15'] = "=IFERROR(LOOKUP(2,1/('Dziennik'!C:C<>\"Błędna formuła\"),'Dziennik'!C:C), \"Brak danych\")"
    ws_settings['B15'].font = formula_font
    ws_settings['B15'].fill = formula_fill

    input_cells_settings = ['B2', 'B3', 'B4', 'B9', 'B11', 'B12']
    for cell in input_cells_settings:
        ws_settings[cell].fill = input_fill
        ws_settings[cell].value = 0

    ws_settings['B2'].value = 1800
    ws_settings['B3'].value = 200
    ws_settings['B4'].value = 300
    ws_settings['B9'].value = 500
    ws_settings['B11'].value = 2.0
    ws_settings['B12'].value = 0.25
    ws_settings['C12'] = "(Wpisz 0.25 dla 25%)"
    ws_settings['C12'].font = info_font

    ws_settings.column_dimensions['A'].width = 30
    ws_settings.column_dimensions['B'].width = 15
    ws_settings.column_dimensions['C'].width = 25


    # --- 2. Zakładka: Dziennik (ZAKTUALIZOWANA) ---
    ws_log = wb.create_sheet("Dziennik")
    print("Tworzę zaktualizowaną zakładkę [Dziennik]...")

    headers_log = [
        # Ogólne
        "Data", "Tydzień",
        # Fizjologia (Rano)
        "Waga (kg)", "Waga (śr. 7-dniowa)", "RHR", "HRV", "Sen (h)", "Jakość snu (1-5)", 
        "Samopoczucie (1-5)",
        # Trening (Wpisywane)
        "Trening (Kcal)", "Trening (Czas, min)", "Jakość Treningu (1-5)",
        # Dolegliwości (Fizjo)
        "Dolegliwości (Opis)", "Dolegliwości (Ból 1-5)",
        # Obliczenia Kaloryczne (Formuły)
        "TDEE (Szacowane)", "CEL Kcal (na dziś)",
        # Spożycie (Wpisywane)
        "Spożyte Kcal",
        # Bilans (Formuła)
        "Bilans Kcal (dnia)",
        # Obliczenia Makro (Formuły)
        "CEL Białko (g)", "CEL Tłuszcze (g)", "CEL Węgle (g)",
        # Spożycie Makro (Wpisywane)
        "Spożyte Białko (g)", "Spożyte Tłuszcze (g)", "Spożyte Węgle (g)",
        # Inne (Wpisywane)
        "Płyny Spożyte (L)", "Suplementy (Notatka)", "Notatki (Ogólne)"
    ]

    # Zaktualizowana lista kolumn do ręcznego wpisania (żółte tło)
    # (1-based index)
    input_columns = [
        1, 3, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15,
        18, 23, 24, 25, 26, 27, 28
    ]

    # Define columns that should have thick right borders (end of logical sections)
    section_end_columns = [
        2,  # After Tydzień
        8,  # After Jakość snu
        9,  # After Samopoczucie
        12, # After Jakość Treningu
        14, # After Dolegliwości
        16, # After CEL Kcal
        17, # After Spożyte Kcal
        18, # After Bilans
        21, # After CEL Węgle
        24, # After Spożyte Węgle
        25, # After Płyny
        26, # After Suplementy
    ]

    for i, header in enumerate(headers_log, 1):
        cell = ws_log.cell(row=1, column=i)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
        # Apply appropriate border
        border_style = thick_right_border if i in section_end_columns else thin_border
        cell.border = border_style
        
        fill_style = input_fill if i in input_columns else formula_fill
        for r in range(2, 1002):
            cell = ws_log.cell(row=r, column=i)
            cell.fill = fill_style
            cell.border = border_style  # Apply same border style to all cells in column

    # --- ZAKTUALIZOWANE Formuły do Wiersza 2 (dopasowane do aktualnych kolumn) ---
    # Mapowanie kolumn (ważne):
    # A - Data
    # B - Tydzień
    # C - Waga (kg)
    # D - Waga (śr. 7-dniowa)
    # J - Trening (Kcal)
    # O - TDEE (Szacowane)
    # P - CEL Kcal (na dziś)
    # Q - Spożyte Kcal
    # R - Bilans Kcal (dnia)
    # S - CEL Białko (g)
    # T - CEL Tłuszcze (g)
    # U - CEL Węgle (g)
    formulas_log = {
        'B2': '=IF(ISNUMBER(A2), WEEKNUM(A2, 2), "")', # Tydzień
        'D2': '=IF(ISNUMBER(C2), AVERAGE(C2:INDEX(C:C, MAX(2, ROW()-6))), "")', # Waga śr. 7-dniowa
        # TDEE: baza z Ustawienia (B6) + treningowe kcal z kolumny J (Trening (Kcal)) jeśli istnieją
        'O2': '=IF(ISNUMBER(J2), \'Ustawienia i Cele\'!$B$6 + J2, \'Ustawienia i Cele\'!$B$6)', # TDEE
        # CEL Kcal: TDEE - planowany deficyt
        'P2': '=O2 - \'Ustawienia i Cele\'!$B$9',
        # Bilans: jeśli brak wpisanego spożycia (Q2) -> pusty, inaczej Q2 - P2
        'R2': '=IF(ISBLANK(Q2), "", Q2 - P2)',
        # CEL Białko: jeśli brak wagi w C2 -> pusty, inaczej współczynnik z ustawień * waga (C2)
    'S2': '=IF(OR(C2="",C2=0), "", IFERROR(\'Ustawienia i Cele\'!$B$11 * C2, 0))',
        # CEL Tłuszcze: procent z ustawień przeliczony na gramy (kcal tłuszczu = P2 * %Tł) / 9
        'T2': '=IFERROR((P2 * \'Ustawienia i Cele\'!$B$12) / 9, 0)',
        # CEL Węgle: pozostałe kcal po białku i tłuszczu podzielone przez 4
        'U2': '=IFERROR((P2 - (S2*4) - (T2*9)) / 4, 0)'
    }
    
    for cell, formula in formulas_log.items():
        ws_log[cell] = formula
        ws_log[cell].font = formula_font

    # A2 is the input date (left empty by default); subsequent rows auto-fill for 90 days when A2 is filled
    ws_log['A2'] = ""
    ws_log['A2'].fill = input_fill
    ws_log['A2'].number_format = 'yyyy-mm-dd'
    # Fill dates down for the next 89 rows (so total entries 90 including A2)
    for r in range(3, 92):
        # e.g. A3 = IF(ISBLANK($A$2), "", $A$2 + (ROW()-2))
        ws_log[f'A{r}'] = '=IF(ISBLANK($A$2), "", $A$2 + (ROW()-2))'
        ws_log[f'A{r}'].number_format = 'yyyy-mm-dd'
    ws_log['AC2'] = "Wypełnij żółte pola. Szare liczą się same." # Przesunięta notatka
    ws_log['AC2'].font = info_font
    ws_log['AC2'].fill = input_fill

    # Zaktualizowane szerokości kolumn
    column_widths = [
        12, 8,  # Data, Tydzień
        10, 10, 8, 8, 8, 10, # Waga, Waga śr, RHR, HRV, Sen, Jakość snu
        12, # Samopoczucie
        12, 12, 12, # Trening Kcal, Czas, Jakość
        25, 12, # Dolegliwości Opis, Ból
        12, 12, # TDEE, CEL Kcal
        12, # Spożyte Kcal
        12, # Bilans
        12, 12, 12, # Cele Makro
        12, 12, 12, # Spożyte Makro
        10, 20, 30 # Płyny, Suple, Notatki
    ]
    
    for i, width in enumerate(column_widths, 1):
        ws_log.column_dimensions[get_column_letter(i)].width = width

    ws_log.freeze_panes = 'A2'


    # Ensure Excel does a full recalculation on load to avoid stale cached values
    try:
        wb.calculation_properties.fullCalcOnLoad = True
    except Exception:
        # older openpyxl versions may use different attribute names — ignore if not available
        pass

    # --- 3. Zakładka: Dashboard (ZAKTUALIZOWANA) ---
    ws_dash = wb.create_sheet("Dashboard")
    print("Tworzę zaktualizowaną zakładkę [Dashboard]...")

    ws_dash['A1'] = "PODSUMOWANIE TYGODNIOWE"
    ws_dash['A1'].font = Font(bold=True, size=16)
    ws_dash['A3'] = "Wpisz nr tygodnia:"
    ws_dash['A3'].font = Font(bold=True)
    ws_dash['B3'] = datetime.date.today().isocalendar()[1]
    ws_dash['B3'].fill = input_fill
    ws_dash['B3'].font = Font(bold=True)

    summary_headers = {'A5': "Wskaźnik", 'B5': "Średnia / Suma", 'C5': "Komentarz"}
    for cell, value in summary_headers.items():
        ws_dash[cell] = value
        ws_dash[cell].font = header_font
        ws_dash[cell].fill = header_fill
        ws_dash[cell].border = thin_border
        
    summary_data = {
        'A6': "Średnia waga (kg)", 'A7': "Śr. Kcal (Spożyte)",
        'A8': "Śr. Bilans Kcal (dnia)", 'A9': "Łączny Czas Treningu (h)",
        'A10': "Śr. Jakość Snu (1-5)", 'A11': "Śr. Samopoczucie (1-5)",
    }
    for cell, value in summary_data.items():
        ws_dash[cell] = value
        ws_dash[cell].font = Font(bold=True)
        
    # ZAKTUALIZOWANE FORMUŁY (AVERAGEIFS / SUMIFS)
    dashboard_formulas = {
        'B6': "=IFERROR(AVERAGEIFS('Dziennik'!C:C, 'Dziennik'!B:B, $B$3), \"Brak danych\")", # Śr. Waga (kol C)
        'B7': "=IFERROR(AVERAGEIFS('Dziennik'!Q:Q, 'Dziennik'!B:B, $B$3), \"Brak danych\")", # Śr. Kcal Spożyte (kol Q)
        'B8': "=IFERROR(AVERAGEIFS('Dziennik'!R:R, 'Dziennik'!B:B, $B$3), \"Brak danych\")", # Śr. Bilans Kcal (kol R)
        'B9': "=IFERROR(SUMIFS('Dziennik'!L:L, 'Dziennik'!B:B, $B$3) / 60, \"Brak danych\")", # Suma Czasu Treningu (kol L) / 60 min
        'B10': "=IFERROR(AVERAGEIFS('Dziennik'!H:H, 'Dziennik'!B:B, $B$3), \"Brak danych\")", # Śr. Jakość Snu (kol H)
        'B11': "=IFERROR(AVERAGEIFS('Dziennik'!I:I, 'Dziennik'!B:B, $B$3), \"Brak danych\")", # Śr. Samopoczucie (kol I)
    }
    
    for cell, formula in dashboard_formulas.items():
        ws_dash[cell] = formula
        ws_dash[cell].font = formula_font
        ws_dash[cell].fill = formula_fill

    ws_dash['C16'] = "INSTRUKCJA DO WYKRESÓW"
    ws_dash['C16'].font = Font(bold=True)
    ws_dash['C17'] = "1. Przejdź do zakładki 'Dziennik'.\n"\
                     "2. Zaznacz kolumny (np. 'Data' i 'Waga (śr. 7-dniowa)').\n"\
                     "3. Wybierz Wstawianie -> Wykres.\n"\
                     "4. Wytnij (Ctrl+X) i wklej (Ctrl+V) go tutaj."
    ws_dash['C17'].alignment = Alignment(wrap_text=True, vertical="top")
    ws_dash['C17'].font = info_font
    ws_dash.row_dimensions[17].height = 70
    
    ws_dash.column_dimensions['A'].width = 30
    ws_dash.column_dimensions['B'].width = 20
    ws_dash.column_dimensions['C'].width = 50


    # --- 4. Zakładka: Źródła CHO (produkty) ---
    ws_cho = wb.create_sheet("Źródła CHO")
    print("Tworzę zakładkę [Źródła CHO] z listą produktów i obliczeniami...")

    cho_headers = [
        "Nazwa produktu", "Porcja (g)", "CHO / 100g (g)", "kcal / 100g",
        "CHO w porcji (g)", "kcal w porcji", "Typ (żel/baton/napój/inne)", "Uwagi"
    ]

    for i, h in enumerate(cho_headers, 1):
        c = ws_cho.cell(row=1, column=i)
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    # Mark input columns (Porcja, CHO/100g, kcal/100g) as editable (yellow)
    # Mark input columns (Porcja, CHO/100g, kcal/100g) as editable (yellow) for many rows
    for r in range(2, 1002):
        ws_cho.cell(row=r, column=2).fill = input_fill
        ws_cho.cell(row=r, column=3).fill = input_fill
        ws_cho.cell(row=r, column=4).fill = input_fill

    # Template formulas in row 2
    # CHO in portion = (Porcja * CHO/100g) / 100
    ws_cho['E2'] = '=IF(OR(B2="",C2=""), "", B2 * C2 / 100)'
    ws_cho['F2'] = '=IF(OR(B2="",D2=""), "", B2 * D2 / 100)'
    ws_cho['E2'].font = formula_font
    ws_cho['F2'].font = formula_font
    ws_cho['E2'].fill = formula_fill
    ws_cho['F2'].fill = formula_fill

    # Example rows (you can edit or remove these) - expanded list
    sample = [
        ("Banan", 100, 23, 89, "owoc", "Klasyczny wybór przed/po treningu"),
        ("Żel energetyczny", 40, 60, 240, "żel", "Szybka dawka CHO podczas wysiłku"),
        ("Rodzynki", 30, 79, 299, "suszone owoce", "Gęste źródło cukrów, wygodne w transporcie"),
        ("Daktyle (Medjool)", 24, 75, 277, "suszone owoce", "Szybkie i skoncentrowane źródło energii"),
        ("Miód", 15, 82, 304, "płynne", "Szybkie cukry, łatwe do dodania do napoju"),
        ("Chleb biały (kromka)", 30, 49, 265, "pieczywo", "Łatwo dostępne źródło węglowodanów"),
        ("Wafelek ryżowy", 9, 85, 387, "wafelek", "Niska gęstość energetyczna; szybkie węgle"),
        ("Baton energetyczny (np. Clif)", 68, 48, 400, "baton", "Kombinacja węgli i tłuszczu — dłuższe uwalnianie energii"),
        ("Napój izotoniczny (Gatorade)", 250, 6.9, 26, "napój", "Płynne źródło CHO i elektrolitów podczas treningu"),
        ("Żelki (gummy)", 40, 72.5, 475, "słodycze", "Szybkie cukry, dobre w krótkich wysiłkach"),
        # Dodane: maltodekstryna i fruktoza oraz ich mieszanka w stosunku 1:0.8
        ("Maltodekstryna", 30, 100, 400, "proszek", "Czyste węglowodany, szybko dostępne"),
        ("Fruktoza", 30, 100, 399, "cukier", "Wolniej metabolizowana niż glukoza; stosować z umiarem"),
        ("Mieszanka MD:FR (1:0.8)", 30, 100, 399, "mieszanka", "Maltodekstryna:Fruktoza w stosunku 1:0.8 — szybkie uzupełnienie CHO")
    ]
    for r, item in enumerate(sample, start=2):
        ws_cho.cell(row=r, column=1).value = item[0]
        ws_cho.cell(row=r, column=2).value = item[1]
        ws_cho.cell(row=r, column=3).value = item[2]
        ws_cho.cell(row=r, column=4).value = item[3]
        ws_cho.cell(row=r, column=7).value = item[4]
        ws_cho.cell(row=r, column=8).value = item[5]
        # apply formula for CHO and kcal in portion for sample rows
        ws_cho.cell(row=r, column=5).value = f'=IF(OR(B{r}="",C{r}=""), "", B{r} * C{r} / 100)'
        ws_cho.cell(row=r, column=6).value = f'=IF(OR(B{r}="",D{r}=""), "", B{r} * D{r} / 100)'
        ws_cho.cell(row=r, column=5).font = formula_font
        ws_cho.cell(row=r, column=6).font = formula_font
        ws_cho.cell(row=r, column=5).fill = formula_fill
        ws_cho.cell(row=r, column=6).fill = formula_fill

    # Column widths and formatting
    # Increase widths for 'Typ' and 'Uwagi' so text fits on one line
    widths = [30, 14, 14, 14, 16, 14, 20, 40]
    for i, w in enumerate(widths, 1):
        ws_cho.column_dimensions[get_column_letter(i)].width = w

    ws_cho.freeze_panes = 'A2'


    # --- Zapisywanie pliku ---
    file_name = "kombajn_triathlonisty_v2.xlsx"
    wb.save(file_name)
    
    print("---------------------------------------------------------")
    print(f"GOTOWE! 🚀")
    print(f"Plik '{file_name}' został stworzony.")
    print("---------------------------------------------------------")
    print("\nJak zacząć:")
    print("1. Otwórz plik i idź do [Ustawienia i Cele].")
    print("2. Idź do [Dziennika]. ŻÓŁTE pola wypełniasz ręcznie.")
    print("3. SZARE pola liczą się same. Przeciągnij formuły z wiersza 2 w dół.")

except ImportError:
    print("[BŁĄD] Nie znaleziono biblioteki 'openpyxl'.")
    print("Uruchom w terminalu: pip install openpyxl")
except PermissionError:
    print(f"[BŁĄD] Nie mogę zapisać pliku. Może masz już otwarty plik '{file_name}'?")
except Exception as e:
    print(f"Wystąpił nieoczekiwany błąd: {e}")